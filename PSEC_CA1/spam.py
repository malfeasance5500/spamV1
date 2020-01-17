import openpyxl,re,datetime, hashlib
# the spam class , which is the app
class SPAM():
    # defining some vars that are to be initialised when the spam class is initialised
    def __init__(self):
        self.items = {} #to store the menu for the day
        self.wb = openpyxl.load_workbook("items.xlsx") #this is the excel file that stores the menu
        self.sheet = None #sheet is a var that will access the sheets in the excel file, set as none as we require the day to get the sheet
        self.menu = {"1":self.getMenu, "2":self.searchMenu, "3":self.displayCart, "4":self.checkOut} #functions stored in a dict so that the function can be called after the input, without need for if else
        self.orderAmount={}# orderAmount is a var that holds key:amt
        self.total=0 #holds the total money to pay
        self.today = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]  #used to check for the day of the week
        self.userList = openpyxl.load_workbook("users.xlsx").active #this is excel sheet containing user name and pw, may hash the passwords?
        self.users = {}
        self.user = "anonymous"
        # init the user list and admin lists
        self.match = []
    

    # function that runs once when the app is started, gets the day of the week, updates the menu and starts up the menu 
    def greet(self): 
       print("==================\nWelcome to SPAM:\n==================\n")
       self.getDay()
       self.updateMenu()
       self.login()
       self.showMenu()
       
    
    # function for the login system of the app
    def login(self):
        # loo through the excel file to get the user name and passwords
        self.userList = openpyxl.load_workbook("users.xlsx").active
        for i in range(1,self.userList.max_row + 1):
            self.users[ self.userList["A"+str(i)].value] = self.userList["B"+str(i)].value
        # code block to validate the input of the user's choice to login or not
        again=True
        while again:
            self.loginChoice = input("Would you like to login(l) or skip(s) or create an account(c): ")
            self.loginChoice = self.loginChoice.strip()
            if self.loginChoice.lower()=="s":
                print("Login skipped")
                again=False
                # if user wants to skip the login, we go straight to the menu
            elif self.loginChoice.lower() =="c":
                self.createUser()
            elif self.loginChoice.lower()=="l":
                again=False
                print("Commencing login\n")
                # actual login of the thing
                self.username = input("Please enter your username: ")
                while len(self.username)==0:
                    self.username = input("Please enter a valid username")
                self.password = input("Please enter your password: ")
                while len(self.password)==0:
                    self.password = input("Please enter a valid password: ")

                #check to see if the username matches anything
                if self.username in list(self.users.keys()):
                    # the passwords in the excel file are hashed, so we can get security and check 
                    # as hash is non reversible, the passwords cannot be reverse engineered
                    if hashlib.md5(self.password.encode()).hexdigest() == self.users[self.username]:
                        print("Login successful.")
                        print(f"Welcome, {self.username}")
                    else:
                        print("Invalid credentials, logging in as anonymous")
                        print("\nWelcome, anonymous\n")
                        
                else:
                    print("Invalid credentials, logging in as anonymous")
                    print("\nWelcome, anonymous\n")
            else:
                print("Please give a valid input!")
        
    def createUser(self):
        username = input("Please input your username: ")
        while len(username)<1:
            print("Please input a proper username")
            username = input("Please input your username: ")
        password = input("Please input your password: ")
        while password=="":
            print("Please enter a proper password")
            password = input("Please input your password: ")
        passwordConfirm = input("Please enter your password again: ")
        if password == passwordConfirm:
            # self.userList["A"+ str(self.userList.max_row +1) ] = str( hashlib.md5(password.encode()).hexdigest())
            workbook = openpyxl.load_workbook("users.xlsx")
            sheet = workbook.active
            sheet["A"+str(self.userList.max_row+1)] = username
            sheet["B"+str(self.userList.max_row+1)] = str( hashlib.md5(password.encode()).hexdigest())

            # Save the spreadsheet
            workbook.save("users.xlsx")
            self.userList = openpyxl.load_workbook("users.xlsx").active
            for i in range(1,self.userList.max_row + 1):
                self.users[ self.userList["A"+str(i)].value] = self.userList["B"+str(i)].value
            # self.login()
        else:
            print("The password does not match")
            # self.login()
    
    # as seen in the name, this function checks for the date, which will allow the program to get the menu for the day from the excel spread sheet
    def getDay(self):
        time = datetime.datetime.today()
        self.today = self.today[ time.weekday() ]
        self.sheet = self.wb[self.today]
    
    # menu function, shows the 4 options available to users
    def showMenu(self):
        # code chunk to validate the input to make sure its a number and it fits the range
        again =True
        while again:
            self.menuChoice = input("1. Display Today's Menu\n2. Search Menu\n3. Display Cart\n4. Check Out\n\nPlease input your choice of action (ENTER to exit): ")
            self.menuReg = re.compile(r"[1-4]") #regex pattern to make sure users input is 1-4
            if re.match(self.menuReg, self.menuChoice)!=None and len(self.menuChoice)==1:
                again=False
                # if the check passes, i execute a function 
                self.menu[self.menuChoice]()

            # if user enters, i exit the app
            elif len(self.menuChoice)==0:
                again=False
                self.exitSPAM()
                
            else:
                again =True
                print("\nInvalid input! \n")
                # at this point, when again is false, the user choices will be sent to my functions'

    # function to update the menu
    def updateMenu(self):
        # self.keys is the list of the items
        self.match=[]
        self.keys = [] 
        # for loop to iterate over the rows of the spreadsheet to get the items and prices, to add to self.items and self.keys
        for i in range(1,self.sheet.max_row + 1):
            self.items[self.sheet["A"+str(i)].value] = self.sheet["B"+str(i)].value
            self.keys.append(self.sheet["A"+str(i)].value)   
            self.match.append(self.sheet["A"+str(i)].value)   

    # function to show users the menu
    def getMenu(self):
        # first function
        print("Menu for today:\n==============")    
        # longest will get the length of the longest key in self.items
        longest = len((max(list(self.items.keys() ), key=len )) )
        # for loop to print the menu out, ljust is used to properly format the menau nicely
        # for i in range(len(self.items)):
        #     print(f"{i+1}. { (self.keys[i]).ljust(longest+3, ' ')} :           ${self.items[self.keys[i]]:.2f}")
        print("\n\n")
        self.updateMenu()
        self.orderFood()
        self.showMenu()

    # function to let users search the menu, casing of input does not matter
    def searchMenu(self):
        # second function

        # while code block to repeat input until the input is not just enter, input cannot be empty
        again =True
        while again:
            self.foodChoice = input("\n\nPlease input food to search: ")
            self.foodChoice = self.foodChoice.strip()
            # if user doesnt input anything, he is asked to input sth
            if len(self.foodChoice)==0:
                print("Please provide input.")
            else:
                again=False
            
        # code to see if the input matches anything in the keys
        # the matching items are stored in this var
        self.match = [] 
        # for loop to loop throught the items to check if the user input matches any of the foods
        for i in range(len(self.keys)):
            if self.foodChoice.lower() in self.keys[i].lower():
                self.match.append(self.keys[i])
        # if nothing matches, print that msg if not, we show them the results
        if len(self.match)==0:
            print(f"Sorry, we do not serve {self.foodChoice}")
        else:
            self.orderFood()
        self.showMenu()

    def orderFood(self):
        longest = len((max(self.match,key=len)))
        for i in range(len(self.match)):
            print(f"{i+1}. { (self.match[i]).ljust(longest+3, ' ')} :           ${self.items[self.match[i]]:.2f}")
        print("\n")

        # code block to proceed to ask user if he wants to order

        again =True
        # regex pattern to make sure that input is 1- wtv the amt of items is
        orderReg = re.compile(rf"[1-{len(self.match)}]")
        while again:
            order = input(f"Enter the dish 1-{len(self.match)} that you would like to order, 0 to stop: ")
            # validation for the ordering  
            if order=="0":
                again=False
                print("You have ordered:\n")
                if len(self.orderAmount)!=0:
                    longest = len((max(list(self.orderAmount.keys() ), key=len)))
                    for i in range(len(list(self.orderAmount.keys()))):
                        print(f"{i+1}. {(list(self.orderAmount.keys())[i]).ljust(longest+3, ' ')} : { ( str(self.orderAmount[list(self.orderAmount.keys())[i] ]) ).ljust(longest+3, ' ') }")
                else:
                    print("You have ordered nothing")
                print("\n\n")
            elif len(order)!=1 or orderReg.match(order)==None :
                print("\nInvalid input!")
            else:
                # get the index of the food ordered, in the self.match list
                buy = int(orderReg.search(order).group(0)) -1 
                check = True
                while check:
                    # ask user how much of the item he wants, inputting 0 will remove it from the cart
                    amount = input(f"How many of {self.match[buy]} would you like to order: ")
                    if amount.isnumeric()==False  or len(amount)==0:
                        print("Please enter a valid amount")
                    elif amount=="0":
                        self.orderAmount.pop( self.match[buy] ,None)
                        check=False
                    else:
                        print(f"{amount} {self.match[buy]} added to cart ")
                        self.orderAmount[self.match[buy]] = int(amount)
                        check=False

    def displayCart(self):
        # third function
        if len(self.orderAmount)!=0:
            print("In your cart is: \n")
            longest = len((max(list( self.orderAmount.keys()),key=len )))
            for i in range(len(list(self.orderAmount.keys()))):
                print(f"{i+1}. {(list(self.orderAmount.keys())[i]).ljust(longest+3, ' ')} : { ( str(self.orderAmount[list(self.orderAmount.keys())[i] ]) ).ljust(longest+3, ' ') } ")
        else:
            print("Your shopping cart is empty.")
        print("\n\n") 

        self.showMenu()

    def checkOut(self):
        # fourth function, checkout
        print("\nPls check your order:\n")
        if len(self.orderAmount)!=0:
            longest = len((max(list(self.orderAmount.keys()),key=len)))
            for i in range(len(list(self.orderAmount.keys()))):
                print(f"{i+1}. {(list(self.orderAmount.keys())[i]).ljust(longest+3, ' ')} : { ( str(self.orderAmount[list(self.orderAmount.keys())[i] ]) ) } x ${self.items[list(self.orderAmount.keys())[i]]:.2f} ")
                self.total+= (( (self.orderAmount[list(self.orderAmount.keys())[i] ]) ) * self.items[list(self.orderAmount.keys())[i]])
            # getting user approval to quit the SPAM 
            quit = input("Do you wish to continue shopping(n) or proceed to payment(y) : ")
            while quit.lower()!= "y" and quit.lower()!="n":
                print("Please enter a valid input")
                quit = input("Do you wish to continue shopping(n) or proceed to payment(y) : ")
            if quit.lower() =="n":
                self.showMenu()
            else:

                print(f"Thank you for using SPAM. Please pay a total of: ${self.total:.2f} ")
                input("Press Enter to continue...")
                self.exitSPAM()
        else:
            print("Your shopping cart is empty.")
            self.showMenu()
        

    def exitSPAM(self):
        print("Thank you for you patronage!")

spam = SPAM()
spam.greet()

